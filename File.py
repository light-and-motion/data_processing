import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import (ScatterChart, Reference, Series)
import numpy as np

class File(object): 
    def __init__(self, mapped_settings, general_settings, output_data, output_name): 
        self.mapped_settings = mapped_settings
        self.general_settings = general_settings
        self.output_data = output_data
        self.output_name = output_name
    
    def make_file(self, choice): 
        """ Determines if an Excel, JPEG, or PDF file will be generated"""

        if (pd.isnull(choice) or choice.upper() == 'YES'): 
            return True
        return False

class ChartFile(File): 
    def make_chart(self):
        """Returns a list that indicates whether there will be a chart and if so, which columns will serve as the x-axis 
        and y-axis 

        Parameters: 
        axis (series): Indicates which CSV columns will serve as the x-axis and the y-axis

        Returns:
        List: 
            a) If first element is False, no chart will be generated
            b) If first element is True, second element will be a one-element series of the column that will serve as the x-axis 
                and third element will be a series of the column(s) that will serve as the y-axis
        """ 
        axis = self.mapped_settings.get_column('Axis')
        if (axis.dropna().empty or not ((axis == 'x').any() or (axis == 'X').any()) or not ((axis == 'y').any() or (axis == 'Y').any())):  
            return False
        return True
    
    def get_x_axis(self): 
        x_axis = None
        if (self.make_chart()): 
            axis = self.mapped_settings.get_column('Axis')
            x_axis = axis.loc[(axis == 'x') | (axis == 'X')]
        return x_axis
    
    def get_y_axis(self): 
        y_axis = None
        if (self.make_chart()): 
            axis = self.mapped_settings.get_column('Axis')
            y_axis = axis.loc[(axis == 'y') | (axis == 'Y')]
        return y_axis        
    
    def chart_title(self, new_titles, chart_title, x_axis_index, y_axis_indices):
        """Returns the chart title. 

        If no title is given, then the chart title will default to the format '[All] y-axis vs x-axis'
        
        Parameters: 
        new_titles (series): New titles of the processed CSV columns 
        chart_title (series): Contain a manually given chart title or NaN
        x_axis_row (series): Index location of the column to serve as the x_axis
        y_axis_row (series): Index location(s) of the column(s) to serve as the y-axis 
        """
        
        # Note: A column with 'NaNs' is not considered empty
        if (chart_title.dropna().empty): 
            title = ''
            for i in range(y_axis_indices.size-1): 
                title += new_titles.loc[y_axis_indices[i]] + ", "
            title += new_titles.loc[y_axis_indices[y_axis_indices.size-1]] + " vs " + new_titles.loc[x_axis_index]
        else: 
            title = chart_title.loc[0]
        return title


class ExcelFile(ChartFile):

    def output_excel(self):
        if (self.make_file(self.general_settings.get_column('Excel').loc[0])): 
            
            # Create workbook to hold output data 
            wb = self.create_plotted_workbook()

            self.process_data(wb)

            if (self.make_chart()): 
               self.create_chart(wb)
            
            wb.save(self.output_name + '.xlsx')
            
    def create_plotted_workbook(self): 
        """Returns an empty Excel workbook of the data to be plotted with the title of the
        default worksheet labeled as "Output Data."
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'Output Data'
        return wb
    
    def process_data(self, wb):
        """Returns a Workbook object where the input data is mapped into the desired columns in the output Excel file 

        Parameters: 
        wb (Workbook object): Store results of data processing 
        df (dataframe): Data to be mapped 
        config_df (dataframe): 'Mapped Settings' of the configuration file  
        output_col_letters (series): Output column letters 
        """
        
        new_titles = new_titles = self.mapped_settings.get_column('Title')
        output_numbers = self.mapped_settings.get_column('Output')

        # Grab active Worksheet
        ws = wb.active

        # Read in all the data 
        for j in range(output_numbers.size): 
            self.read_in_values(ws, new_titles.iloc[j], output_numbers.iloc[j])
        return wb
    
    
    def read_in_values(self, ws, title, col_num):
        """Reads in the data of 1 to-be processed CSV column into the Excel workbook 

        Parameters: 
        wb (workbook): Store the results of the data processing 
        title (str): New titles of the processed CSV columns   
        col_num (int): Column number the data is being read into   
        """ 

        header = ws.cell(row=1, column = col_num) 
        header.value = title
        header.font = Font(bold=True)
        #col_index = title_input
        
       
        # Indices: i retrieves the data in the column 
        #          cellRow ensures that the data is being mapped to the current cell in the Excel worksheet
        cellRow = 2 
        i = 0

        # Determine the size of the current column 
        size = self.output_data[title].size

        while (i < size):   
            ws.cell(row = cellRow, column = col_num).value = self.output_data.loc[i,title]
            cellRow += 1
            i += 1
    
    def create_chart(self,wb): 
        """Creates a chart sheet of the plotted data in the output Excel workbook

        Parameters: 
        wb (workbook): Excel workbook of the mapped data 
        """

        ws = wb.active
        
        outputs = self.mapped_settings.get_column('Output')  
        new_titles = self.mapped_settings.get_column('Title')
        chart_title = self.general_settings.get_column('Chart Title')
        row_size = self.output_data[new_titles.loc[0]].size # get the row number of the last cell 
        
        # Create a ScatterChart chart sheet 
        cs = wb.create_chartsheet()
        chart = ScatterChart()

        # Store the index of the x_axis column in mapped_settings
        x_axis_index= self.get_x_axis().index[0] 

        # Store the column number where the x_axis is located 
        x = Reference(ws, min_col=outputs.loc[x_axis_index], min_row = 2, max_row = row_size)
               
        # Plot as many y-axis as indicated in the configuration file 
        y_axis_indices = self.get_y_axis().index
        for row in y_axis_indices: 
            y = Reference(ws, min_col = outputs.loc[row], min_row = 2, max_row = row_size)
            s = Series(y,x,title=new_titles.loc[row])
            chart.append(s)
        
        chart.x_axis.title = new_titles.loc[x_axis_index]
        
        # Situate x-axis below negative numbers 
        chart.x_axis.tickLblPos = "low"

        # Create the chart legend or set the y-axis label 
        self.chart_legend(chart, new_titles.loc[y_axis_indices[0]], y_axis_indices) 
    
        # Title the chart
        chart.title = self.chart_title(new_titles, chart_title, x_axis_index, y_axis_indices)

        # Set grid lines on or off. 
        self.grid_lines(chart)

        # Set the scaling limits of the x and y axis
        self.chart_scaling(chart)

        cs.add_chart(chart)

    
    def chart_legend(self, chart, y_label, y_axis_indices):
        """Returns True if a legend is needed, False otherwise. """
        if (len(y_axis_indices) == 1):
            chart.y_axis.title = y_label
            chart.legend = None  
        else: 
            pass # a legend is the default 
            
    '''
    def chart_title(self, new_titles, chart_title, x_axis_index, y_axis_indices):
        """Returns the chart title. 

        If no title is given, then the chart title will default to the format '[All] y-axis vs x-axis'
        
        Parameters: 
        new_titles (series): New titles of the processed CSV columns 
        chart_title (series): Contain a manually given chart title or NaN
        x_axis_row (series): Index location of the column to serve as the x_axis
        y_axis_row (series): Index location(s) of the column(s) to serve as the y-axis 
        """
        
        # Note: A column with 'NaNs' is not considered empty
        if (chart_title.dropna().empty): 
            title = ''
            for i in range(y_axis_indices.size-1): 
                title += new_titles.loc[y_axis_indices[i]] + ", "
            title += new_titles.loc[y_axis_indices[y_axis_indices.size-1]] + " vs " + new_titles.loc[x_axis_index]
        else: 
            title = chart_title.loc[0]
        return title
    '''
    def grid_lines(self, chart): 
        """Returns True if grid lines will be on chart, False otherwise"""
        
        isGridLinesOn = self.general_settings.get_column('Grid Lines').loc[0]
        if (not pd.isnull(isGridLinesOn) and isGridLinesOn.upper() == 'NO'): 
            chart.x_axis.majorGridlines = None 
            chart.y_axis.majorGridlines = None

        return False 
    
    def chart_scaling(self,chart): 
        """Returns a list of the settings for the minimum and maximum of the x and y axis"""
        
        x_min = self.general_settings.get_column('X Min').loc[0]
        x_max = self.general_settings.get_column('X Max').loc[0]
        y_min = self.general_settings.get_column('Y Min').loc[0]
        y_max = self.general_settings.get_column('Y Max').loc[0]

        if (not pd.isnull(x_min)): 
            chart.x_axis.scaling.min = x_min 
        if (not pd.isnull(x_max)): 
            chart.x_axis.scaling.max = x_max
        if (not pd.isnull(y_min)): 
            chart.y_axis.scaling.min = y_min 
        if (not pd.isnull(y_max)): 
            chart.y_axis.scaling.max = y_max
      
     
