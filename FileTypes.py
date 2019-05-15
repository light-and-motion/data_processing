from File import (File, ChartFile)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import (ScatterChart, Reference, Series)
from datetime import (datetime, time)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()
import pdfkit
from PyPDF2 import PdfFileReader, PdfFileWriter
import os
import numpy as np
import pandas as pd

class ExcelFile(ChartFile):

    def output(self):
        if (self.make_file(self.general_settings.get_column('Excel').loc[0])): 
            
            # Create workbook to hold output data 
            wb = self._create_plotted_workbook()

            self._process_data(wb)

            if (self.make_chart()): 
               self._create_chart(wb)
            
            wb.save(self.output_name + '.xlsx')
            
    def _create_plotted_workbook(self): 
        """Returns an empty Excel workbook of the data to be plotted with the title of the
        default worksheet labeled as "Output Data."
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'Output Data'
        return wb
    
    def _process_data(self, wb):
        """Returns a Workbook object where the input data is mapped into the desired columns in the output Excel file 

        Parameters: 
        wb (Workbook object): Store results of data processing 
        df (dataframe): Data to be mapped 
        config_df (dataframe): 'Mapped Settings' of the configuration file  
        output_col_letters (series): Output column letters 
        """
        
        new_titles = new_titles = self.mapped_settings.get_column('Title')
        output_numbers = self.mapped_settings.get_column('Output')

        # Grab active Worksheet
        ws = wb.active

        # Read in all the data 
        for j in range(output_numbers.size): 
            self._read_in_values(ws, new_titles.iloc[j], output_numbers.iloc[j])
        return wb
    
    
    def _read_in_values(self, ws, title, col_num):
        """Reads in the data of 1 to-be processed CSV column into the Excel workbook 

        Parameters: 
        wb (workbook): Store the results of the data processing 
        title (str): New titles of the processed CSV columns   
        col_num (int): Column number the data is being read into   
        """ 

        header = ws.cell(row=1, column = col_num) 
        header.value = title
        header.font = Font(bold=True)
        #col_index = title_input
        
       
        # Indices: i retrieves the data in the column 
        #          cellRow ensures that the data is being mapped to the current cell in the Excel worksheet
        cellRow = 2 
        i = 0

        # Determine the size of the current column 
        size = self.output_data[title].size

        while (i < size):   
            ws.cell(row = cellRow, column = col_num).value = self.output_data.loc[i,title]
            cellRow += 1
            i += 1
    
    def _create_chart(self,wb): 
        """Creates a chart sheet of the plotted data in the output Excel workbook

        Parameters: 
        wb (workbook): Excel workbook of the mapped data 
        """

        ws = wb.active
        
        outputs = self.mapped_settings.get_column('Output')  
        new_titles = self.mapped_settings.get_column('Title')
        chart_title = self.general_settings.get_column('Chart Title')
        row_size = self.output_data[new_titles.loc[0]].size # get the row number of the last cell 
        
        # Create a ScatterChart chart sheet 
        cs = wb.create_chartsheet()
        chart = ScatterChart()

        # Store the index of the x_axis column in mapped_settings
        x_axis_index= self.get_x_axis().index[0] 

        # Store the column number where the x_axis is located 
        x = Reference(ws, min_col=outputs.loc[x_axis_index], min_row = 2, max_row = row_size)
               
        # Plot as many y-axis as indicated in the configuration file 
        y_axis_indices = self.get_y_axis().index
        for row in y_axis_indices: 
            y = Reference(ws, min_col = outputs.loc[row], min_row = 2, max_row = row_size)
            s = Series(y,x,title=new_titles.loc[row])
            chart.append(s)
        
        chart.x_axis.title = new_titles.loc[x_axis_index]
        
        # Situate x-axis below negative numbers 
        chart.x_axis.tickLblPos = "low"

        # Create the chart legend or set the y-axis label 
        self._chart_legend(chart, new_titles.loc[y_axis_indices[0]], y_axis_indices) 
    
        # Title the chart
        chart.title = self.get_chart_title(new_titles, chart_title, x_axis_index, y_axis_indices)

        # Set grid lines on or off. 
        self._grid_lines(chart)

        # Set the scaling limits of the x and y axis
        self._chart_scaling(chart)

        cs.add_chart(chart)

    
    def _chart_legend(self, chart, y_label, y_axis_indices):
        """Returns True if a legend is needed, False otherwise. """
        if (len(y_axis_indices) == 1):
            chart.y_axis.title = y_label
            chart.legend = None  
        else: 
            pass # a legend is the default 
            
    def _grid_lines(self, chart): 
        """Returns True if grid lines will be on chart, False otherwise"""
        
        isGridLinesOn = self.general_settings.get_column('Grid Lines').loc[0]
        if (not pd.isnull(isGridLinesOn) and isGridLinesOn.upper() == 'NO'): 
            chart.x_axis.majorGridlines = None 
            chart.y_axis.majorGridlines = None

        return False 
    
    def _chart_scaling(self,chart): 
        """Returns a list of the settings for the minimum and maximum of the x and y axis"""
        
        x_min = self.general_settings.get_column('X Min').loc[0]
        x_max = self.general_settings.get_column('X Max').loc[0]
        y_min = self.general_settings.get_column('Y Min').loc[0]
        y_max = self.general_settings.get_column('Y Max').loc[0]

        if (not pd.isnull(x_min)): 
            chart.x_axis.scaling.min = x_min 
        if (not pd.isnull(x_max)): 
            chart.x_axis.scaling.max = x_max
        if (not pd.isnull(y_min)): 
            chart.y_axis.scaling.min = y_min 
        if (not pd.isnull(y_max)): 
            chart.y_axis.scaling.max = y_max

class JPEGFile(ChartFile): 
    def output(self): 
        jpeg_choice = self.make_file(self.general_settings.get_column('JPEG').loc[0])
        pdf_choice = self.make_file(self.general_settings.get_column('PDF').loc[0])
        if ((jpeg_choice or pdf_choice) and self.make_chart()):
            self._make(jpeg_choice, pdf_choice) 
    
    def _make(self, jpeg_choice, pdf_choice):  
        """Produces a JPG and/or PDF file of a matplotlib chart

        Parameters: 
        jpeg_choice (bool): True if chart will be saved as JPEG, False otherwise
        pdf_choice (bool): True if chart will be saved as PDF, False otherwise
        """
        
        new_titles = self.mapped_settings.get_column('Title')
        chart_title = self.general_settings.get_column('Chart Title')
        x_axis_index = self.get_x_axis().index[0]
        y_axis_indices = self.get_y_axis().index
        is_x_elapsed_time = False
        is_y_elapsed_time = False

        # Plot multiple lines on a single chart. 
        # As matplotlib does not allow timedelta objects to be directly set as an axis, must convert to a 
        # datetime object to plot on chart. 
        x_axis = self.output_data[new_titles[x_axis_index]].dropna() #mapping_df[new_titles[x_axis_row.index[0]]].dropna() 
        if (not (pd.isnull(self.mapped_settings.get_column('Time Unit').loc[x_axis_index]))):
            is_x_elapsed_time = True
            x_axis = pd.Series(self._convert_timedelta_to_datetime(x_axis))
            
        
        #TODO: ReminderIf there are multiple y-axes, their dtypes have to be the same! 
        #TODO: Why does the data range have to be the same even columns that will not be plotted
        #fig, ax = plt.subplots(1,1)
        fig, ax = plt.subplots(1) #figure(1)
        for y_axis_index in y_axis_indices: 
            y_axis_title = new_titles[y_axis_index]
            y_axis = self.output_data[y_axis_title].dropna()
            if (not pd.isnull(self.mapped_settings.get_column('Time Unit').loc[y_axis_index])): 
                is_y_elapsed_time = True
                y_axis = pd.Series(self._convert_timedelta_to_datetime(y_axis))
            ax.plot(x_axis, y_axis, label = new_titles.iloc[y_axis_index])
       
        
        # Elapsed time formatting
        if (is_x_elapsed_time): 
            self._format_x_date(fig, ax)
        
        if (is_y_elapsed_time): 
            self._format_y_date(fig, ax)
        
        # Set the x-label and the y-label/legend
        self._chart_legend(ax, x_axis_index, y_axis_indices)
        
        # Set the title 
        title = self.get_chart_title(new_titles, chart_title, x_axis_index, y_axis_indices)
        ax.set_title(title)
        
        # Set gridlines 
        self._grid_lines(ax)   

        # Chart scaling 
        self._chart_scaling(ax)
        
        # Save charts in stated formats
        
        if (jpeg_choice): 
            plt.savefig(self.output_name + '.jpeg', bbox_inches = 'tight')
        
        if (pdf_choice): 
            plt.savefig(self.output_name + '_chart' + '.pdf', bbox_inches = 'tight') 
        return fig
        

    #TODO: Faster way to acomplish this https://stackoverflow.com/questions/48294332/plot-datetime-timedelta-using-matplotlib-and-python
    def _convert_timedelta_to_datetime(self,timedelta_series): 
        """Takes in a Series that contains timedelta objects and returns a Series that contains datetime objects"""
        
        # Convert 'timedelta_series' to type str 
        timedelta_str_series = timedelta_series.astype(str)
        #print(timedelta_str_series)

        # Split 'timedelta_str_series' using the space delimiter and store the results into a list
        timedelta_str_list = [time.split() for time in timedelta_str_series]
        
        # Retrieve the 'time' portion of 'timedelta_str_list' and store into another list  
        time_str_list = [time[2] for time in timedelta_str_list]
    
        # Split 'time_str_list' using '.' delimiter and store results back into 'time_str_list'  
        time_str_list = [time.split('.') for time in time_str_list]

        # Retrieve the '%H:%M:%S' formatted time and store results back into list 
        time_str_list = [time[0] for time in time_str_list]

        # Convert 'time_str_list' into a series and turn each element into a datetime.time() object. 
        # Store in a new list. 
        time_str_series = pd.Series(time_str_list)
        time_obj = [datetime.strptime(time_str, '%H:%M:%S').time() for time_str in time_str_series]
        x_axis = [ datetime.combine(datetime.now(), time) for time in time_obj]
        
        return x_axis

    def _chart_legend(self, ax, x_axis_index, y_axis_indices): 
        # Set the labels and/or legend of the chart
        
        new_titles = self.mapped_settings.get_column('Title') 
        ax.set_xlabel(new_titles[x_axis_index])
        if (len(y_axis_indices) > 1):
            ax.legend(loc='upper left', bbox_to_anchor =(1.05,1))
        else: 
            ax.set_ylabel(new_titles[y_axis_indices[0]])

    def _grid_lines(self, ax): 
        isGridLinesOn = self.general_settings.get_column('Grid Lines') 
        if (pd.isnull(isGridLinesOn.loc[0]) or isGridLinesOn.loc[0].upper() == 'YES'): 
            ax.grid(b = True)
    
    def _format_x_date(self, fig, ax): 
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            fig.autofmt_xdate()
    
    def _format_y_date(self, fig, ax): 
            ax.yaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    
    def _chart_scaling(self, ax): 
        x_min = self.general_settings.get_column('X Min').loc[0]
        x_max = self.general_settings.get_column('X Max').loc[0]
        y_min = self.general_settings.get_column('Y Min').loc[0]
        y_max = self.general_settings.get_column('Y Max').loc[0]

        if (not pd.isnull(x_min)): 
            ax.set_xlim(left = x_min)
        if (not pd.isnull(x_max)): 
            ax.set_xlim(right = x_max)
        if (not pd.isnull(y_min)): 
            ax.set_ylim(bottom = y_min)
        if (not pd.isnull(y_max)): 
            ax.set_ylim(top = y_max)
        

class PDFFile (ChartFile): 

    def output(self): 
        pdf_choice = self.make_file(self.general_settings.get_column('PDF').loc[0])
        if (pdf_choice): 
            self._make()
    
    def _make(self): 
        """Generates a pdf of the processed results 

        Parameters: 
        output_name (str): Name PDF will be saved as 
        mapping_data_df (dataframe): CSV columns to be processed
        create_chart (bool): True if a chart will be generated in the PDF, False if not 
        """
        
        # Get the file path of the wkhtmltopdf executable 
        #config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        df_file = os.getcwd() + '\\' + self.output_name + '.pdf'
        
        # Replace NaN values with empty strings so the empty data cells do not look like they hold any values in the PDF file 
        mapping_df = self.output_data.fillna('')
        
        # Convert datetime into strings so 0 days portion doesn't show up in PDF
        total_time_cols = self.mapped_settings.get_col('Time Units')
    
        mapping_df['Date/Time'] = [date[-8:] for date in mapping_df['Date/Time'].astype(str)]
        
        # If the PDF file is to contain a chart, then merge the dataframe and chart PDF into a single PDF. 
        # Otherwise, just save the dataframe PDF as is. 
        if (not self.make_chart()): 
            pdfkit.from_string(mapping_df.to_html(), df_file)
    
        else:  
            df_file = os.getcwd() + '\\' + self.output_name + '_table.pdf'
            pdfkit.from_string(mapping_df.to_html(), df_file)
            paths = [os.getcwd() + '\\' + self.output_name + '_chart.pdf' ,df_file]
            self._merge_pdfs(paths)

    def _merge_pdfs(self,paths): 
        """Merges two PDFs into a single PDF 
        
        Source: https://realpython.com/pdf-python/

        Parameters: 
        paths (list): File paths of the PDFs to be merged  
        output_name (str): Name PDF will be saved as  
        """
        
        pdf_writer = PdfFileWriter()

        for path in paths: 
            pdf_reader = PdfFileReader(path)
            for page in range(pdf_reader.getNumPages()):
                pdf_writer.addPage(pdf_reader.getPage(page))
        
        with open(self.output_name + '.pdf', 'wb') as out: 
            pdf_writer.write(out)
        
        # Delete merged files 
        os.remove(paths[0])
        os.remove(paths[1])

class TXTFile(File): 
    def output(self): 
        text_choice = self.make_file(self.general_settings.get_column('TXT').loc[0])
        if(text_choice): 
            self._make()

    def _make(self): 
        """Generates a text file of the processed results"""
        #print(mapping_df.head())
        mapping_array = self.output_data.to_numpy()
        my_fmt = self._get_format()

        #TODO: tab delimiter looks weird on txt file
        np.savetxt(self.output_name + '.txt', mapping_array, fmt = my_fmt, delimiter='\t', header = '\t'.join([str(column) for column in self.output_data.columns]), comments='')


    def _get_format(self): 
        dtypes = self.output_data.dtypes
        fmt = []
        for i in range(len(dtypes)):
            type = dtypes[i] 
            if (type == np.int64): 
                fmt.append('%d')
            # Parse floats as strings because %f truncates the length of the (very long!) floats 
            else: 
                fmt.append('%s')
        return fmt

      
     
