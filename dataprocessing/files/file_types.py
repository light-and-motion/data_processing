from datetime import (datetime, time)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np
import pandas as pd
from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import (ScatterChart, Reference, Series)
from PyPDF2 import PdfFileReader, PdfFileWriter
from .file import (File, ChartFile)

class ExcelFile(ChartFile):
    """
    Extends ChartFile to output an Excel file of the processed CSV results. 
    """
    
    def get_name(self): 
        return self.output_name + '.xlsx'

    def will_output_file(self): 
        if (self.make_file(self.general_settings.get_column('Excel').loc[0])): 
            return True 
        return False 
    
    def output(self) -> None:
        """ Outputs an Excel file"""

        if (self.will_output_file()): 
            
            # Create workbook to hold output data 
            wb = self._create_plotted_workbook()

            self._process_data(wb)

            if (self.make_chart()): 
               self._create_chart(wb)
            
            wb.save(self.get_name())
            
    def _create_plotted_workbook(self) -> Workbook: 
        """Returns an empty Excel workbook that will later hold the processed CSV data.
        
        Helper function to output(). 
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'Output Data'
        return wb
    
    def _process_data(self, wb):
        """Map the input data into the specified columns in the Excel workbook.   
        
        Helper function to output(). 

        Parameters
        ----------
        wb : Workbook 
            Empty workbook to store input data 

        Returns
        ------- 
        Workbook
            Filled with extracted CSV data 
        """
        
        new_titles = new_titles = self.mapped_settings.get_column('Title')
        output_numbers = self.mapped_settings.get_column('Output')

        # Grab active Worksheet
        ws = wb.active

        # Write in all the data from the extracted CSV columns  
        for j in range(output_numbers.size): 
            self._write_in_values(ws, new_titles.iloc[j], output_numbers.iloc[j])
        return wb
    
    
    def _write_in_values(self, ws, label, col_num):
        """Writes the data of a CSV column into an Excel worksheet. 

        Helper function to _process_data(). 

        Parameters
        ----------
        ws : Worksheet
            Worksheet that CSV column is writing into 
        label : str
            Column labels of the CSV columns in the new worksheet      
        col_num : int
            Column number the input data is being written into in the new worksheet   

        Returns
        ------ 
        None 
        """ 

        header = ws.cell(row=1, column = col_num) 
        header.value = label
        header.font = Font(bold=True)
        #col_index = title_input
        
       
        # Indices: i retrieves the data in the column 
        #          cellRow ensures that the data is being mapped to the current cell in the Excel worksheet
        cellRow = 2 
        i = 0

        # Determine the size of the current column 
        size = self.output_data[label].size

        while (i < size):   
            ws.cell(row = cellRow, column = col_num).value = self.output_data.loc[i,label]
            cellRow += 1
            i += 1
    
    def _create_chart(self,wb): 
        """Creates a chart sheet of the processed CSV data in the Excel workbook. 

        Parameters
        ----------
        wb (workbook): Excel workbook of the mapped data 

        Returns 
        -------
        None 
        """

        ws = wb.active
        
        outputs = self.mapped_settings.get_column('Output')  
        new_titles = self.mapped_settings.get_column('Title')
        chart_title = self.general_settings.get_column('Chart Title')
        row_size = self.output_data[new_titles.loc[0]].size # Get the row number of the last cell 
        
        # Create a ScatterChart chart sheet 
        cs = wb.create_chartsheet()
        chart = ScatterChart()

        # Store the row indices of the x-axis and y-axis column labels in the configuration 
        # file mapped_settings.  
        x_axis_index= self.get_x_axis().index[0] 
        y_axis_indices = self.get_y_axis().index

        # Set x-axis 
        x = Reference(ws, min_col=outputs.loc[x_axis_index], min_row = 2, max_row = row_size)
               
        # Plot multiple graphs in a single chart  
        for row in y_axis_indices: 
            y = Reference(ws, min_col = outputs.loc[row], min_row = 2, max_row = row_size)
            s = Series(y,x,title=new_titles.loc[row])
            chart.append(s)
        
        # Set the x-axis label
        chart.x_axis.title = new_titles.loc[x_axis_index]
        
        # Situate x-axis below negative numbers 
        chart.x_axis.tickLblPos = "low"

        # Create the chart legend or set the y-axis label 
        self._chart_legend(chart, new_titles.loc[y_axis_indices[0]], y_axis_indices) 
    
        # Title the chart
        chart.title = self.get_chart_title(new_titles, chart_title, x_axis_index, y_axis_indices)

        # Set grid lines on or off. 
        self._grid_lines(chart)

        # Set the scaling limits of the x and y axis
        self._chart_scaling(chart)

        # Add chart to the workbook 
        cs.add_chart(chart)

    
    def _chart_legend(self, chart, y_label, y_axis_indices) -> None:
        """Sets the chart legend. If there is only 1 plotted line, set the y-label."""

        if (len(y_axis_indices) == 1):
            chart.y_axis.title = y_label
            chart.legend = None  
        else: 
            pass # a legend is the default 
            
    def _grid_lines(self, chart) -> None: 
        """Set the chart grid lines on or off."""
        
        isGridLinesOn = self.general_settings.get_column('Grid Lines').loc[0]
        if (not pd.isnull(isGridLinesOn) and isGridLinesOn.upper() == 'NO'): 
            chart.x_axis.majorGridlines = None 
            chart.y_axis.majorGridlines = None
    
    def _chart_scaling(self,chart) -> None: 
        """Set the scales on the x and y axis"""
        
        x_min = self.general_settings.get_column('X Min').loc[0]
        x_max = self.general_settings.get_column('X Max').loc[0]
        y_min = self.general_settings.get_column('Y Min').loc[0]
        y_max = self.general_settings.get_column('Y Max').loc[0]

        if (not pd.isnull(x_min)): 
            chart.x_axis.scaling.min = x_min 
        if (not pd.isnull(x_max)): 
            chart.x_axis.scaling.max = x_max
        if (not pd.isnull(y_min)): 
            chart.y_axis.scaling.min = y_min 
        if (not pd.isnull(y_max)): 
            chart.y_axis.scaling.max = y_max

class JPEGFile(ChartFile): 
    """
    Extends ChartFile to output a JPEG file fo the processed CSV results.
    """

    def get_name(self): 
        return self.output_name + '.jpeg'
    
    def will_output_file(self): 
        if (self.make_file(self.general_settings.get_column('JPEG').loc[0]) 
            and self.make_chart()): 
            return True 
        return False 

    def output(self) -> None: 
        """Outputs an JPEG file."""

        jpeg_choice = self.make_file(self.general_settings.get_column('JPEG').loc[0])
        pdf_choice = self.make_file(self.general_settings.get_column('PDF').loc[0])
        if ((jpeg_choice or pdf_choice) and self.make_chart()):
            self._make(jpeg_choice, pdf_choice) 
    
    def _make(self, jpeg_choice, pdf_choice):  
        """Produces a JPG and/or PDF file of a matplotlib chart. 
        
        Helper function to output(). 

        Parameters
        ----------
        jpeg_choice : bool 
            True if chart will be saved as JPEG, False otherwise
        pdf_choice : bool
            True if chart will be saved as PDF, False otherwise

        Returns 
        -------
        None
        """
        
        new_titles = self.mapped_settings.get_column('Title')
        chart_title = self.general_settings.get_column('Chart Title')
        x_axis_index = self.get_x_axis().index[0]
        y_axis_indices = self.get_y_axis().index
        is_x_elapsed_time = False
        is_y_elapsed_time = False

        # As matplotlib does not allow timedelta objects to be directly set as an axis, we need to convert 
        # the timedelta objects into datetime objects to plot them on a chart.
        x_axis = self.output_data[new_titles[x_axis_index]].dropna() 
        if (not (pd.isnull(self.mapped_settings.get_column('Time Unit').loc[x_axis_index]))):
            is_x_elapsed_time = True
            x_axis = pd.Series(self._convert_timedelta_to_datetime(x_axis))
            
        
        # Reminder: If there are multiple plots, their dtypes have to be the same! 
        fig = plt.figure(figsize = (8.5, 11))
        ax = plt.subplot()

        # Plot multiple graphs in a single chart. 
        for y_axis_index in y_axis_indices: 
            y_axis_title = new_titles[y_axis_index]
            y_axis = self.output_data[y_axis_title].dropna()
            if (not pd.isnull(self.mapped_settings.get_column('Time Unit').loc[y_axis_index])): 
                is_y_elapsed_time = True
                y_axis = pd.Series(self._convert_timedelta_to_datetime(y_axis))
            ax.plot(x_axis, y_axis, label = new_titles.iloc[y_axis_index])
       
        
        # Elapsed time formatting
        if (is_x_elapsed_time): 
            self._format_x_date(fig, ax)
        
        if (is_y_elapsed_time): 
            self._format_y_date(fig, ax)
        
        # Set the x-label and the y-label/legend
        self._chart_legend(ax, x_axis_index, y_axis_indices)
        
        # Set the title 
        title = self.get_chart_title(new_titles, chart_title, x_axis_index, y_axis_indices)
        ax.set_title(title)
        
        # Set gridlines 
        self._grid_lines(ax)   

        # Chart scaling 
        self._chart_scaling(ax)
        
        # Set page number 
        fig.text(4.25/8.5, 0.5/11., '1', ha='center', fontsize=12)

        # Save charts in stated formats
        if (jpeg_choice): 
            plt.savefig(self.get_name(), bbox_inches = 'tight')
        
        if (pdf_choice): 
            plt.savefig(self.output_name + '_chart' + '.pdf', bbox_inches = 'tight') 
    
        

    #TODO: Faster way to acomplish this https://stackoverflow.com/questions/48294332/plot-datetime-timedelta-using-matplotlib-and-python
    def _convert_timedelta_to_datetime(self,timedelta_series): 
        """
        Converts a series with timedelta objects into a series with datetime objects. 

        Parameters 
        ----------
        timedelta_series: pd.Series
            Contains timedelta objects

        Returns
        -------
        pd.Series
            Contains datetime objects 
        """
        
        # Convert 'timedelta_series' to type str 
        timedelta_str_series = timedelta_series.astype(str)

        # Split 'timedelta_str_series' using the space delimiter and store the results into a list
        timedelta_str_list = [time.split() for time in timedelta_str_series]
        
        # Retrieve the 'time' portion of 'timedelta_str_list' and store into another list  
        time_str_list = [time[2] for time in timedelta_str_list]
    
        # Split 'time_str_list' using '.' delimiter and store results back into 'time_str_list'  
        time_str_list = [time.split('.') for time in time_str_list]

        # Retrieve the '%H:%M:%S' formatted time and store results back into list 
        time_str_list = [time[0] for time in time_str_list]

        # Convert 'time_str_list' into a series and turn each element into a datetime object. 
        # Store in a new list. 
        time_str_series = pd.Series(time_str_list)
        time_obj = [datetime.strptime(time_str, '%H:%M:%S').time() for time_str in time_str_series]
        x_axis = [ datetime.combine(datetime.now(), time) for time in time_obj]
        
        return x_axis

    def _chart_legend(self, ax, x_axis_index, y_axis_indices) -> None: 
        """Sets the chart legend. If there is only 1 plotted line, set the y-label."""
        
        new_titles = self.mapped_settings.get_column('Title') 
        ax.set_xlabel(new_titles[x_axis_index])
        if (len(y_axis_indices) > 1):
            ax.legend(loc='upper left', bbox_to_anchor =(1.05,1))
        else: 
            ax.set_ylabel(new_titles[y_axis_indices[0]])

    def _grid_lines(self, ax) -> None: 
        """Set the chart grid lines on or off."""

        isGridLinesOn = self.general_settings.get_column('Grid Lines') 
        if (pd.isnull(isGridLinesOn.loc[0]) or isGridLinesOn.loc[0].upper() == 'YES'): 
            ax.grid(b = True)
    
    def _format_x_date(self, fig, ax) -> None: 
        """Format a datetime x-axis so that the labels do not show the day, only the time"""
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))

        # Position the labels so they do not overlap with each other. 
        fig.autofmt_xdate()
    
    def _format_y_date(self, fig, ax) -> None: 
        """Format a datetime y-axis so that the labels do not show the day, only the time."""

        ax.yaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    
    def _chart_scaling(self, ax) -> None: 
        """Set the scales on the x and y axis"""

        x_min = self.general_settings.get_column('X Min').loc[0]
        x_max = self.general_settings.get_column('X Max').loc[0]
        y_min = self.general_settings.get_column('Y Min').loc[0]
        y_max = self.general_settings.get_column('Y Max').loc[0]

        if (not pd.isnull(x_min)): 
            ax.set_xlim(left = x_min)
        if (not pd.isnull(x_max)): 
            ax.set_xlim(right = x_max)
        if (not pd.isnull(y_min)): 
            ax.set_ylim(bottom = y_min)
        if (not pd.isnull(y_max)): 
            ax.set_ylim(top = y_max)
        

class PDFFile (ChartFile): 
    """
    Extends ChartFile to output a PDF file of the processed CSV results. 
    """

    def get_name(self): 
        return self.output_name + '.pdf'

    def will_output_file(self): 
        if (self.make_file(self.general_settings.get_column('PDF').loc[0])): 
            return True
        return False 

    def output(self) -> None: 
        """Outputs a PDF file."""

        if (self.will_output_file()): 
            self._make()
    
    def _make(self) -> None: 
        """Generates a PDF of the processed results.  

        Helper function to output(). 
        """
        
        # Grab the file path of the 
        df_file = os.getcwd() + '\\' + self.output_name + '.pdf'
        
        # Replace NaN values with empty strings so the empty data cells do not 
        # look like they hold any values in the PDF file. 
        # Convert datetime into strings so 0 days portion doesn't show up in PDF
        mapping_df = self.get_str_timedelta()

        # Output the table in a PDF. 
        self._make_table(mapping_df)
        
        # If the PDF file is to contain a chart, then merge the dataframe and chart PDF into a single PDF. 
        if (self.make_chart()):   
            df_file = os.getcwd() + '\\' + self.output_name + '_table.pdf'
            #pdfkit.from_string(mapping_df.to_html(), df_file)
            paths = [os.getcwd() + '\\' + self.output_name + '_chart.pdf' ,df_file]
            self._merge_pdfs(paths)

    # def timedelta_to_string(self,mapping_df): 
        
    #     indices = self.mapped_settings.get_column('Time Unit').dropna().index
    #     for i in indices: 
    #         label = self.mapped_settings.get_column('Title').loc[i]
    #         mapping_df[label] = [time[7:15] for time in mapping_df[label].astype(str)]
    #     return mapping_df

    def _make_table(self, mapping_df): 
        """
        Renders the output dataframe in PDF format. 

        Source: https://stackoverflow.com/questions/19726663/how-to-save-the-pandas-dataframe-series-data-as-a-figure

        Parameters: 
        mapping_df (pd.DataFrame): Output dataframe with NaNs replaced with whitespace

        Returns: 
        None 
        """

        if (self.make_chart()): 
            pp = PdfPages(self.output_name + '_table.pdf')
        else: 
            pp = PdfPages(self.get_name())
        total_rows, total_cols = mapping_df.shape

        rows_per_page = 40 # Assign a page cut off length
        rows_printed = 0

        if (self.make_chart()): 
            page_number = 2
        else: 
            page_number = 1

        while (total_rows >0): 

            # Put the table on a correctly sized figure    
            fig = plt.figure(figsize=(8.5,11))
            ax = plt.subplot()
            ax.axis('off')
            
            #try: 
             #   matplotlib_tab = pd.plotting.table(ax, mapping_df.iloc[rows_printed:rows_printed+rows_per_page], loc='upper center', colWidths=[0.2, 0.2, 0.2])    
            #except IndexError:
            matplotlib_tab = pd.plotting.table(ax, mapping_df.iloc[rows_printed:rows_printed+rows_per_page], loc='upper center')

            # Style the cells  
            table_props=matplotlib_tab.properties()
            table_cells=table_props['child_artists'] # I have no clue why child_artists works
            for cell in table_cells:
                    cell.set_height(0.024)
                    cell.set_fontsize(14)
                    
            # Add a header and footer with page number 
            fig.text(4.25/8.5, 10.5/11., "Table", ha='center', fontsize=12)
            fig.text(4.25/8.5, 0.5/11., str(page_number), ha='center', fontsize=12)

            pp.savefig()
            plt.close()

            #Update variables
            rows_printed += rows_per_page
            total_rows -= rows_per_page
            page_number+=1

        pp.close()


    def _merge_pdfs(self,paths): 
        """Merges two PDFs into a single PDF 
        
        Source: https://realpython.com/pdf-python/

        Parameters
        ---------- 
        paths : list
            File paths of the PDFs to be merged  

        Returns 
        --------
        None    
        """
        
        pdf_writer = PdfFileWriter()
        for path in paths: 
            pdf_reader = PdfFileReader(path)
            for page in range(pdf_reader.getNumPages()):
                pdf_writer.addPage(pdf_reader.getPage(page))
        
        with open(self.get_name(), 'wb') as out: 
            pdf_writer.write(out)
        
        # Delete merged files 
        os.remove(paths[0])
        os.remove(paths[1])

class TXTFile(File): 
    """
    Extends File to output a text file of the processed CSV results. 
    """

    def get_name(self): 
        return self.output_name + '.txt'

    def will_output_file(self): 
        if (self.make_file(self.general_settings.get_column('TXT').loc[0])):
            return True
        return False 

    def output(self): 
        """Outputs a text file."""

        if(self.will_output_file()): 
            self._make()

    def _make(self) -> None: 
        """Generates a text file of the processed results. 

        Helper function to output(). Converts the DataFrame into a NumPy array 
        and uses Numpy functionality to save the array as a text file. 
        """
        mapping_df = self.get_str_timedelta()
        mapping_array = mapping_df.to_numpy()
        my_fmt = self._get_format()

        #FIXME: tab delimiter looks weird on txt file
        np.savetxt(self.get_name(), mapping_array, fmt = my_fmt, delimiter='\t', header = '\t'.join([str(column) for column in self.output_data.columns]), comments='')


    def _get_format(self) -> list: 
        """Determine how to the data will be formatted in the text file."""

        dtypes = self.output_data.dtypes
        fmt = []
        for i in range(len(dtypes)):
            type = dtypes[i] 
            if (type == np.int64): 
                fmt.append('%d')
            else:  # Parse floats as strings because %f truncates the length of the (very long!) floats. 
                fmt.append('%s')
        return fmt

      
     
