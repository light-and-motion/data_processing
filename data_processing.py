import numpy as np
import pandas as pd
from datetime import (datetime, timedelta, time)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import (ScatterChart, Reference, Series)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import pdfkit
from PyPDF2 import PdfFileReader, PdfFileWriter
import os
from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()
class Data_Processing: 
    """
    A class used to process the data of a CSV file.  

    Attributes:
    config_file (str): The name of the configuration file 
    input_csv (str): The name of the CSV file to be proccesed 
    output_name (str): The name the output file(s) will be saved as  
    """    
    
    def __init__(self, config_file, input_csv, output_name): 
        """
        Parameters: 
        config_file (str): The name of the configuration file 
        input_csv (str): The name of the CSV file to be proccesed 
        output_name (str): The name the output file(s) will be saved as 
        """

        self.config_file = config_file
        self.input_csv = input_csv 
        self.output_name = output_name
    
    
    def create_csv_dataframe(self, file, config_df_2): 
        """Returns a dataframe of the CSV file 

        Parameters: 
        file (str): Name of CSV file to be processed
        config_df_2 (dataframe): 'General Settings' of the configuration file 
        """

        start_ser = config_df_2['Start Row']
        stop_ser = config_df_2['Stop Row']
        skip_ser = config_df_2['Skip Row']
        transpose_ser = config_df_2['Transpose']
        
        # Default values         
        startLine = 0
        stopLine = None
        skipLine = None
        transpose = "NO"


        if (not start_ser.dropna().empty): 
            startLine = start_ser.loc[0]-1
        if (not stop_ser.dropna().empty): 
            stopLine = stop_ser.loc[0]-startLine-1
        if (not skip_ser.dropna().empty):
            skipLine = skip_ser.loc[0]
        if (not transpose_ser.dropna().empty): 
            transpose = transpose_ser.loc[0]

        # Read the CSV into the dataframe         
        df = self.read_csv_type(file, startLine, stopLine, skipLine, transpose) 
  
        if (transpose.upper() == 'YES'): 
            #TODO: Ask if N/A marker is to be kept or if the cells that contain it be empty instead. 
            df = self.transpose_df(df, startLine, skipLine)

        ## Minus 1 is added because old column of transposed df has been dropped 
        if (not skipLine == None): 
            df.drop(skipLine-startLine-1, inplace=True)

        # Get rid of columns with all whitespace 
        df = df.dropna('columns', how='all')

        # Reset index to start at 0 after dropping column(s)
        df = df.reset_index(drop=True)
        
        # Search for the columns that have a PM time. (Note: Excel convert str times with a PM time into 24 hour time). 
        # For example, 1:27 PM is converted into 13:27.  
        datetime_str_columns = self.search_for_pm_times(df)
            
        # Format the PM time columns into 12 hour time format. 
        [self.date_parser(df[column_name]) for column_name in datetime_str_columns]        
        
        # As the transpose() function converts the dtypes of the transposed dataframe all into objects when the original dtypes 
        # were mixed, the while loop converts the numeric values back into their proper dtype

        """Returns a series with float datatypes, where the cells with empty Strings are dropped
        
        Parameters: 
        series (series): A series with empty Strings in cells which prevent the series from being of float datatype
        """

        # Iterates through all the columns of the dataframe and converts the numeric values back into their proper datatype

        # Used in particular for: 
        #   a) transposed df: As the transpose() function converts the dtypes of the transposed dataframe all into 
        #                     objects when the original dtypes were mixed, the while loop converts the 
        #                     numeric values back into their proper dtype. 
        #   b) Columns with empty Strings: The empty String values conver the entire column into an object dtype. By dropping the empty strings, 
        #                     the column can be converted to numeric type.
        for column in df: 
            df[column].replace('', np.nan, inplace=True)
            df[column] = df[column].dropna()
            df[column] = pd.to_numeric(df[column], errors = 'ignore')
        
        return df

    def read_csv_type(self, file, startLine, stopLine, skipLine, transpose):
        """Returns the prototype dataframe of the CSV file 
        
        Parameters: 
        file (str):  Name of CSV file to be processed
        startLine (int): Row to begin processing CSV file
        stopLine (int): Row to stop processing CSV file
        skipLine (int): Line you want to skip when processing CSV file. Has to be between startLine and stopLine
        transpose (str): Determines whether df is to be transposed 
        """ 

        # Read to the very end of the file  
        if (stopLine == None and transpose == 'NO'): 
            return pd.read_csv(file + '.csv', 
                            skiprows= startLine, 
                            keep_default_na = False, 
                            encoding = 'ISO-8859-1')

        # Stop reading before the end of a file 
        elif (transpose.upper() == 'NO'):
            return pd.read_csv(file + '.csv', 
                            skiprows= startLine, 
                            nrows = stopLine, 
                            keep_default_na = False, 
                            encoding = 'ISO-8859-1')

        # Read to the very end and transpose df 
        elif (stopLine == None and transpose.upper() == 'YES'): 
            return pd.read_csv(file + '.csv', 
                            header = None, 
                            index_col= None, 
                            skiprows= startLine, 
                            keep_default_na = False, 
                            encoding = 'ISO-8859-1') 
        # Stop reading before the end and transpose df 
        else: 
            return pd.read_csv(file + '.csv', 
                            header = None, 
                            index_col= None, 
                            skiprows= startLine,
                            nrows = stopLine,  
                            keep_default_na = False, 
                            encoding = 'ISO-8859-1') 

    def transpose_df(self, df, startLine, skipLine): 
        """
        Returns a transposed df 

        Parameters: 
        df (df): df that is to be transposed
        startLine (int): Row to begin processing CSV file
        skipLine (int): Line you want to skip when processing CSV file. Has to be between startLine and stopLine
        """
        # Logic to set the actual columns and indices in the transposed data
        df = df.transpose()
        df.rename(df.iloc[0], axis = 'columns', inplace = True)
        df.drop(0, inplace = True)

        #TODO: Research how dataframes are passed into functions. (pass by value or reference)
                
        # As the transpose() function converts the dtypes of the transposed dataframe all into objects when the original dtypes 
        # were mixed, the while loop converts the numeric values back into their proper dtype
        
        return df
    
    def search_for_pm_times(self, df):
        """ Returns a list that stores all the titles of the columns that contain PM times"""

        # Search the first row of every column in the dataframe, convert every value to strptime(%-m/%-d/%Y %H:%M:%S) or 
        # strptime(%-m/%d/%Y %H:%M). If a ValueError is NOT returned, then add the title of the column to the list. 
        datetime_str_column = []
        
        for column in df: 
            series = df[column]
            try: 
                for format in ('%m/%d/%Y %H:%M', '%m/%d/%Y %H:%M:%S'): 
                    datetime_str = str(series.loc[0])
                    datetime.strptime(datetime_str, format)
                    datetime_str_column.append(column)
            except ValueError: 
                pass 
        return datetime_str_column
        
    
    def date_parser(self, datetime_str_series):
        """ Returns a series with its PM times formatted to look like AM times 
        
        Parameters: 
        datetime_str_series (series): Contains PM times that needs to be reformatted  
        """

        # Filter out AM time; they do not need to undergo re-formatting 
        datetime_str_pm = datetime_str_series[~datetime_str_series.str.contains('AM')]
    
        # Return a date format equal to the AM times 
        for datetime_str in datetime_str_pm:  
            datetime_str_list = datetime_str.split()
            date = datetime_str_list[0]
            time = datetime_str_list[1]
            time_list = time.split(':')
            hours = str(int(time_list[0])-12)
            minutes = time_list[1]
            if (len(time_list) == 3): 
                seconds = time_list[2]
            else: 
                seconds = '00'
            new_str = date + ' ' + hours + ':' +  minutes + ':' +  seconds + ' PM'   
            datetime_str_series.replace(datetime_str, new_str, inplace=True)
        return datetime_str_pm

    def create_excel_dataframe(self, file, sheet): 
        """Returns a dataframe of an Excel file 

        Used to store configuration data into a dataframe. 

        Parameters: 
        file (str): Name of Excel file  
        sheet (str): Name of Excel sheet from 'file'  
        """

        df = pd.read_excel(file + '.xlsx', sheet_name = sheet, dtype = {'Title': str})
        return df    

    def create_raw_Excelbook(self, data_df):  
        """Returns an Excel workbook that holds the CSV file in a dataframe

        Parameters: 
        data_df (dataframe): CSV file to be be read into the Excel workbook 
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'Raw Data'
        
        for row in dataframe_to_rows(data_df, index = False, header = True):            
            ws.append(row)
        
        wb.save(self.get_input_csv + '.xlsx')
        return wb

    
    def convert_to_float(self, series): 
        """Returns a series with float datatypes, where the cells with empty Strings are dropped
        
        Parameters: 
        series (series): A series with empty Strings in cells which prevent the series from being of float datatype
        """
        
        series.replace('', np.nan,inplace=True)
        series = series.dropna()
        series = pd.to_numeric(series)#series.astype(float, errors = 'ignore')
        return series

    def create_mapping_dataframe(self, raw_data_df, title_inputs, new_titles, range_inputs, format):
        """Returns a dataframe that contains only the CSV columns that are being processed 

        Parameters: 
        raw_data_df (dataframe): dataframe of CSV file 
        title_inputs (series): Original titles of the processed CSV columns 
        range_inputs (series): Interval of data we want read in each processed CSV column
        """
        
        # Initialize an empty dataframe which will eventually store all mapped values 
        df = pd.DataFrame()
        
        # Find size of dataframe column  
        max_size = raw_data_df.iloc[:,0].size

        # Determine the column with the largest range interval, whose index will be used for the entire dataframe. 
        interval_index = self.largest_range_interval(range_inputs, max_size)
        range_list = self.find_range(range_inputs.loc[interval_index],max_size)
        start = range_list[0]
        end = range_list[1]
        df[title_inputs.loc[interval_index]] = raw_data_df[title_inputs.loc[interval_index]].iloc[start:end].reset_index(drop = True)
        if (not pd.isnull(format.iloc[interval_index]) and type(format.iloc[interval_index]) == np.float64):
            df[title_inputs.loc[interval_index]] = self.round_numbers(df[title_inputs.loc[interval_index]], int(format.iloc[interval_index]))
        df.rename({title_inputs.loc[interval_index]: new_titles.iloc[interval_index]}, axis = 'columns', inplace=True)

        # Drop the rows/columns that have already been used above
        raw_data_df = raw_data_df.drop([title_inputs.loc[interval_index]], axis = 1)
        title_inputs = title_inputs.drop(labels = interval_index).reset_index(drop = True)
        range_inputs = range_inputs.drop(labels = interval_index).reset_index(drop = True)
        format = format.drop(labels = interval_index).reset_index(drop = True)
        new_titles = new_titles.drop(labels = interval_index).reset_index(drop = True)

        # Store all the data to be processed into a dataframe and append each new column to the dataframe 
        for i in range(len(range_inputs)): 
            range_list = self.find_range(range_inputs.loc[i],max_size)
            start = range_list[0]
            end = range_list[1]
            new_series = raw_data_df[title_inputs.loc[i]].iloc[start:end].reset_index(drop = True)

            # Round numbers
            if (not pd.isnull(format.iloc[i]) and type(format.iloc[i]) == np.float64):
                new_series = self.round_numbers(new_series, int(format.iloc[i]))
            df[title_inputs.loc[i]] = new_series
            
            # Rename column titles
            df.rename({title_inputs.loc[i]: new_titles.loc[i]}, axis = 'columns', inplace=True)

        return df

    def round_numbers(self, series, round_to): 
        """ Round numbers in 'series' to the number of decimal places indiciated by 'round_to'"""
        series = series.round(round_to)
        return series
    
    def largest_range_interval(self, range_inputs, max_size):
        """ Returns the row index of the largest range interval  

        Parameters: 
        range_inputs (series): Interval of data we want read in each processed CSV column
        max_size (int) - Size of the CSV column 
        """ 
        
        max_interval = 0
        max_interval_index = 0
        i = 0
        while (i < range_inputs.size): 
            range_list = self.find_range(range_inputs.iloc[i],max_size)
            range_difference = range_list[1]-range_list[0]
            if (range_difference > max_interval): 
                max_interval = range_difference
                max_interval_index = i
            i += 1
            
        return max_interval_index
    
    def find_range(self, current_range, max_size): 
        """Returns a list of the interval of data we want processed in a CSV column.

        The first elment in the list is the starting row index, the second element is the ending row index. 

        Parameters: 
        current_range (float): The interval of the data to be read in 'start:end' format (inclusive)
        max_size (int) - Size of the CSV column 
        """
        
        start = 0 
        end = max_size-1

        # Range is calculated against the row indexes of the Excel worksheet. Thus, the first
        # cell in a column will be located in row 2.   
        if (pd.isnull(current_range)): 
            pass
        else: 
            range_list = current_range.split(':')
            if (range_list[0] == ''):
                end = int(range_list[1])-2
            elif (range_list[1] == ''): 
                start = int(range_list[0])-2
            else: 
                start = int(range_list[0])-2
                end = int(range_list[1])-2
                if (start < 0):
                    return [0, end] 
        return [start,end]
    

    def convert_to_time_object(self, series, unit):
        """Returns a series that contains a str representation of a time object in %H:%M:%S format 

        Parameters: 
        series (Series): The CSV column that is to be converted. Will either contain floats 
                            or a datetime object in %m/%d/%Y %H:%M:%S AM/PM format. 
        unit (str): The unit of time of the CSV column. D = datetime, H = hours, M = minutes, S = seconds 
        """

        series = series.dropna()

        # Creating a copy of 'series' to be iterated over. That way even if two cells have the same datetime, 
        # and both cells are replaced in the original 'series,' the for loop will not break when we iterate over the second copy.  
        series_copy = series.copy()

        # If the time series is a datetime object....
        if (unit.upper() == 'D'): 
            for cur_datetime in series_copy: 
                # Split the datetime object into a list by a space delimiter and store the %H:%M:%S  
                # portion into a variable 
                cur_datetime_list = cur_datetime.split()
                cur_time_list = cur_datetime_list[1].split('.')
                cur_time = cur_time_list[0]
                series.replace(cur_datetime, cur_time, inplace = True)
            
        #If the time series is in hours, minutes, or seconds...
        else:
            for time in series_copy: 
                time_list = self.get_hours_minutes_seconds(time, unit)
                cur_time = str(time_list[0]) + ':' + str(time_list[1]) + ':' + str(time_list[2])
                series.replace(time, cur_time, inplace = True)
        return series

    def get_hours_minutes_seconds(self, time, time_unit):
        """Returns a list that holds the hours, minutes, and seconds of the given time. 

        Takes in a float representation of a a single unit of time, converts it into an integer, and 
        returns a list of the original time in %H:%M:%S format. 

        Parameters: 
        time (float): Given time to be converted
        time_unit: The unit of time that the given time is in
        """
        
        if (time_unit.upper() == 'H'): 
            time = time * 3600
        elif (time_unit.upper() == 'M'): 
            time = time * 60    
        time = int(time)
        hours = time // 3600 
        time = time % 3600
        minutes = time // 60 
        seconds = time % 60
        
        return [hours,minutes,seconds]

    def time_format(self, time_series, start_time): 
        """Returns a series that contains timedelta objects. 

        Takes in a str series in %H:%M:%S format and returns a time series that gives the elapsed time in the same format. 

        Parameters: 
        time_series (series): str representation of time in %H:%M:%S
        """

        time_series_modified = time_series.dropna()
        start_time = pd.to_timedelta(start_time)

        # Iterate through 'time_series_modified' which has the NaN values dropped but replace the str with the timedelta in 
        # the original 'time_series.'
        for current_time in time_series_modified: 
            
            # Convert the str 'current_time' into a timedelta object and find the difference between 'current_time' and 'start_time'
            # to find the elapsed time. Convert the difference between the two times into a str and use the space 
            # delimiter to split it into a list. 
            difference= str(pd.to_timedelta(current_time)-start_time)
            difference_list = difference.split()
            
            # Store the time portion of the str into 'elapsed_time' and use the colon delimiter to split the time into a list 
            elapsed_time = difference_list[2]
            elapsed_time_list = elapsed_time.split(':')
            
            # Convert each element in the list into an integer and use the elements to produce a timedelta object 
            time = timedelta(hours = int(elapsed_time_list[0]), minutes = int(elapsed_time_list[1]), seconds = int(elapsed_time_list[2]))            
        
            # Replace 'current_time' with 'time' in 'time_series'
            time_series.replace(current_time, time, inplace = True)
        return time_series
    
    def create_plotted_workbook(self): 
        """Returns an empty Excel workbook of the data to be plotted with the title of the
        default worksheet labeled as "Output Data."
        """

        wb = Workbook()
        ws = wb.active
        ws.title = 'Output Data'
        return wb

    def convert_columns(self, config_df, col_names):
        """Returns a configuration dataframe, where several mapped settings have been altered. 
        Changes: 
            a) Column letters in 'Input' have been replaced by column titles of the given CSV columns
            b) Column letters in 'Output' have been replaced by column numbers of the given CSV columns 
            c) Empty column titles in the 'title' column have been filled in with original titles of the CSV columns 

        Parameters: 
        config_df (dataframe): 'Mapped Settings' of the configuration file 
        col_names (series): Original titles of the CSV columns 
        """

        config_df['Input Column Numbers'] = config_df['Input'].str.upper()
        config_df['Input'] = self.letter2title(config_df['Input'], col_names)
        self.letter2int(config_df['Output'])
        config_df['Title'] = self.default_titles(config_df['Title'], config_df['Input'])

        return config_df
    
    def letter2title(self, letter_series, names):
        """Returns a series where column letters are converted into column titles.

        Parameters: 
        letter_series (series): Excel column letters
        names (series): CSV column titles 
        """
        
        col_title = []
        indices = self.letter2int(letter_series)
        x = 0
        for col_letter in letter_series: 
            index = indices.loc[x]      
            title = names[index-1]
            col_title.append(title)
            x += 1
        return pd.Series(col_title)
    
    def letter2int(self, letter_series):
        """Returns a series where column letters are being converted into their corresponding column number. 

        Source: https://www.geeksforgeeks.org/find-excel-column-number-column-title/

        Parameters: 
        letter_series (series): Excel column letters
        """
        
        result = 0
        for col_letter in letter_series: 
            result = 0
            for x in col_letter: 
                x = x.upper()
                result *= 26
                result += ord(x) - ord('A') + 1   
            letter_series.replace(col_letter, result, inplace=True)
        return letter_series
    
    def default_titles(self, new_titles, input_titles): 
        """Returns a series where processed CSV columns that are not given a new title in output
        now hold their original CSV column titles. 
        """
        
        x = 0
        for title in new_titles: 
            if (pd.isnull(title)): 
                new_titles.iat[x] = input_titles.iat[x]
            x += 1
        return new_titles

    def process_data(self, wb, df, config_df, output_col_letters):
        """Returns a Workbook object where the input data is mapped into the desired columns in the output Excel file 

        Parameters: 
        wb (Workbook object): Store results of data processing 
        df (dataframe): Data to be mapped 
        config_df (dataframe): 'Mapped Settings' of the configuration file  
        output_col_letters (series): Output column letters 
        """
        
        new_titles = config_df['Title']
        title_inputs = config_df['Input']
        outputs = config_df['Output']

        # Rename the column titles of 'df'
        df.rename(new_titles, axis = 'columns')

        # Grab active Worksheet
        ws = wb.active

        # Read in all the data 
        for j in range(new_titles.size): 
            self.read_in_values(ws, df, new_titles.iloc[j], outputs.iloc[j])
        #self.adjust_column_widths(ws, df, output_col_letters, new_titles)
        return wb

   
    def adjust_column_widths(self, ws, mapping_df, output_col_letters, new_titles):
        """Adjust the column width of the Excel output file
        
        Parameters: 
        ws (worksheet): Worksheet that data is being read into 
        mapping_df (dataframe): CSV columns to be processed  
        output_col_letters (series): Output column letters 
        title_inputs (series): Original titles of the processed CSV columns 
        new_titles (series): New titles of the processed CSV columns 
        """

        i = 0
        series_length = 0
    
        for letters in output_col_letters:
            str_series =  mapping_df[new_titles.iloc[i]].astype(str)

            # Turning a float series into str may suddenly replace values with a datetime-like format
            if (str_series.str.contains('days').loc[0]): 
                series_length= 8
            else: 
                str_series = str_series.map(len)
                series_length = str_series.max()
            
            max_length = max(int(series_length), len(new_titles.iloc[i]))
            ws.column_dimensions[letters].width = max_length
            i += 1

    
    
    def read_in_values(self, ws, mapping_df, new_title, col_num):
        """Reads in the data of 1 to-be processed CSV column into the Excel workbook 

        Parameters: 
        wb (workbook): Store the results of the data processing 
        mapping_df (series): CSV columns to be processed 
        new_title (str): New titles of the processed CSV columns  
        title_inputs (str): Original titles of the processed CSV columns    
        col_num (int): Column number the data is being read into   
        """ 
        
        header = ws.cell(row=1, column = col_num) 
        header.value = new_title
        header.font = Font(bold=True)
        #col_index = title_input
        
       
        # Indices: i retrieves the data in the column 
        #          cellRow ensures that the data is being mapped to the current cell in the Excel worksheet
        cellRow = 2 
        i = 0
        size = mapping_df[new_title].size
        while (i < size):   
            ws.cell(row = cellRow, column = col_num).value = mapping_df.loc[i,new_title]
            cellRow += 1
            i += 1
 
    def make_file(self, choice): 
        """ Determines if an Excel, JPEG, or PDF file will be generated"""

        if (pd.isnull(choice) or choice.upper() == 'YES'): 
            return True
        return False

    def make_chart(self,axis):
        """Returns a list that indicates whether there will be a chart and if so, which columns will serve as the x-axis 
        and y-axis 

        Parameters: 
        axis (series): Indicates which CSV columns will serve as the x-axis and the y-axis

        Returns:
        List: 
            a) If first element is False, no chart will be generated
            b) If first element is True, second element will be a one-element series of the column that will serve as the x-axis 
                and third element will be a series of the column(s) that will serve as the y-axis
        """ 
        
        if (axis.dropna().empty or not ((axis == 'x').any() or (axis == 'X').any()) or not ((axis == 'y').any() or (axis == 'Y').any())):  
            return [False]
        
        x_axis = axis.loc[(axis == 'x') | (axis == 'X')]
        y_axis = axis.loc[(axis == 'y') | (axis == 'Y')]
        return [True, x_axis, y_axis]

    def create_chart(self,wb, mapping_df, x_axis, y_axis, config_df_1, config_df_2): 
        """Creates a chart sheet of the plotted data in the output Excel workbook

        Parameters: 
        wb (workbook): Excel workbook of the mapped data 
        mapping_df (dataframe) - CSV columns to be processed
        x_axis (Series): Indicate which column will serve as the x-axis 
        y_axis (Series): Indicate which column(s) will serve as the y-axis
        config_df_1 (dataframe): 'Mapped Settings' of the configuration file  
        config_df_2 (dataframe): 'General Settings' of the configuration file 
        """

        ws = wb.active
        
        #title_inputs = config_df_1['Input']
        outputs = config_df_1['Output']
        new_titles = config_df_1['Title']
        chart_title = config_df_2['Chart Title']

        row_size = mapping_df[new_titles.loc[0]].size
        
        cs = wb.create_chartsheet()
        chart = ScatterChart()

        # Store the index location of the x-axis value 
        x_axis_row= x_axis.index[0] 

        # Store the column number where the x_axis is located 
        x = Reference(ws, min_col=outputs.loc[x_axis_row], min_row = 2, max_row = row_size)
        
        # Plot as many y-axis as indicated in the configuration file 
        y_axis_rows = y_axis.index
        for row in y_axis_rows: 
            y = Reference(ws, min_col = outputs.loc[row], min_row = 2, max_row = row_size)
            s = Series(y,x,title=new_titles.loc[row])
            chart.append(s)
        
        chart.x_axis.title = new_titles.loc[x_axis_row]
        
        # Situate x-axis below negative numbers 
        chart.x_axis.tickLblPos = "low"

        # Determine whether not there is more than 1 y-axis, which would necessitate the 
        # creation of a legend
        create_legend = self.chart_legend(y_axis_rows) 
        if (not create_legend): 
            chart.y_axis.title = new_titles.loc[y_axis_rows[0]]
            chart.legend = None 
        
        # Title the chart
        chart.title = self.chart_title(new_titles, chart_title, x_axis_row, y_axis_rows)

        # Determine whether grid lines should be on or off. 
        grid_lines = self.grid_lines(config_df_2['Grid Lines'].loc[0])
        if (not grid_lines): 
            
            chart.x_axis.majorGridlines = None 
            chart.y_axis.majorGridlines = None

        # Chart scaling 
        scale = self.chart_scaling(config_df_2['X Min'].loc[0], config_df_2['X Max'].loc[0], config_df_2['Y Min'].loc[0], 
                    config_df_2['Y Max'].loc[0])
        chart.x_axis.scaling.min = scale[0]
        chart.x_axis.scaling.max = scale[1]
        chart.y_axis.scaling.min = scale[2]
        chart.y_axis.scaling.max = scale[3]

        
        cs.add_chart(chart)

    def chart_legend(self, y_axis_rows):
        """Returns True if a legend is needed, False otherwise. """
        
        if (len(y_axis_rows) == 1): 
            return False
        return True
        
    def chart_title(self, new_titles, chart_title, x_axis_row, y_axis_rows):
        """Returns the chart title. 

        If no title is given, then the chart title will default to the format '[All] y-axis vs x-axis'
        
        Parameters: 
        new_titles (series): New titles of the processed CSV columns 
        chart_title (series): Contain a manually given chart title or NaN
        x_axis_row (series): Index location of the column to serve as the x_axis
        y_axis_row (series): Index location(s) of the column(s) to serve as the y-axis 
        """
        
        # Note: A column with 'NaNs' is not considered empty
        if (chart_title.dropna().empty): 
            title = ''
            for i in range(y_axis_rows.size-1): 
                title += new_titles.loc[y_axis_rows[i]] + ", "
            title += new_titles.loc[y_axis_rows[y_axis_rows.size-1]] + " vs " + new_titles.loc[x_axis_row]
        else: 
            title = chart_title.loc[0]
        return title

    def grid_lines(self, choice): 
        """Returns True if grid lines will be on chart, False otherwise"""
        
        if (pd.isnull(choice) or choice.upper() == 'YES'): 
            return True
        return False 

    def chart_scaling(self, x_min, x_max, y_min, y_max): 
        """Returns a list of the settings for the minimum and maximum of the x and y axis"""
        
        x_min_scale = None
        x_max_scale = None
        y_min_scale = None
        y_max_scale = None
        if (not pd.isnull(x_min) and (type(x_min) == np.float64 or type(x_min) == np.int64)): 
            x_min_scale = x_min
        if (not pd.isnull(x_max) and (type(x_min) == np.float64 or type(x_min) == np.int64)): 
            x_max_scale = x_max
        if (not pd.isnull(y_min) and (type(x_min) == np.float64 or type(x_min) == np.int64)): 
            y_min_scale = y_min
        if (not pd.isnull(y_max) and (type(x_min) == np.float64 or type(x_min) == np.int64)): 
            y_max_scale = y_max
        return [x_min_scale, x_max_scale, y_min_scale, y_max_scale]
        
    def make_jpeg(self, mapping_df, x_axis_row, y_axis_row, config_df_1, config_df_2, output_name, jpeg_choice, pdf_choice):  
        """Produces a JPG and/or PDF file of a matplotlib chart

        Parameters: 
        mapping_df (dataframe): CSV columns to be processed 
        x_axis_row (Series): Index location of the column to serve as the x_axis
        y_axis_row (Series): Index location of the column(s) to serve as the y-axis
        config_df_1 (dataframe): 'Mapped Settings' of the configuration file  
        config_df_2 (dataframe): 'General Settings' of the configuration file 
        output_name (str): Name JPG file will be saved as 
        jpeg_choice (bool): True if chart will be saved as JPEG, False otherwise
        pdf_choice (bool): True if chart will be saved as PDF, False otherwise
        """
        
        new_titles = config_df_1['Title']
        title_inputs = config_df_1['Input']
        chart_title = config_df_2['Chart Title']

        
        # Plot multiple lines on a single chart. 
        # As matplotlib does not allow datetime.time objects to be set as an axis, must convert to a 
        # datetime object to plot on chart. 
        x_axis = mapping_df[new_titles[x_axis_row.index[0]]].dropna() 
        x_index = x_axis_row.index[0]
        
        if (not (pd.isnull(config_df_1['Time Unit'].loc[x_index]))):
            x_axis = pd.Series(self.convert_timedelta_to_datetime(x_axis))
            
        fig, ax = plt.subplots(1,1)
        for new_y_index in y_axis_row.index: 
            y_axis_title = new_titles[new_y_index]
            y_axis = mapping_df[y_axis_title]
            if (not pd.isnull(config_df_1['Time Unit'].loc[new_y_index])): 
                y_axis = self.convert_timedelta_to_datetime(y_axis)
            plt.plot(x_axis, y_axis, label = new_titles.iloc[new_y_index])

        # Gives the rows that holds the titles of the columns to be plotted 
        x_axis_rows = x_axis_row.index[0] 
        y_axis_rows = y_axis_row.index 
        
        # Set the labels and/or legend of the chart 
        plt.xlabel(new_titles[x_axis_row.index[0]])
        create_legend = self.chart_legend(y_axis_rows)
        if (create_legend):
            plt.legend(loc='upper left')
        else: 
            plt.ylabel(new_titles[y_axis_row.index[0]])

        # Set the title 
        title = self.chart_title(new_titles, chart_title, x_axis_rows, y_axis_rows)
        plt.title(title)  
        
        # Set gridlines 
        grid_lines = self.grid_lines(config_df_2['Grid Lines'].loc[0])
        if (grid_lines): 
            plt.grid(b = True)
        
        # Date formatter 
        if (not config_df_1['Time Unit'].dropna().empty):
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            fig.autofmt_xdate()

        # Chart scaling 
        scale = self.chart_scaling(config_df_2['X Min'].loc[0], config_df_2['X Max'].loc[0], config_df_2['Y Min'].loc[0], 
                    config_df_2['Y Max'].loc[0])
        plt.axis(scale)

        # Chart scaling 
        scale = self.chart_scaling(config_df_2['X Min'].loc[0], config_df_2['X Max'].loc[0], config_df_2['Y Min'].loc[0], config_df_2['Y Max'].loc[0])
        plt.xlim(scale[0], scale[1])
        plt.ylim(scale[2], scale[3])

        # Save charts in stated formats
        
        if (jpeg_choice): 
            plt.savefig(output_name + '.jpeg')
        
        if (pdf_choice): 
            plt.savefig(output_name + '_chart' + '.pdf') 
        return fig


    def convert_timedelta_to_datetime(self,timedelta_series): 
        """Takes in a Series that contains timedelta objects and returns a Series that contains datetime objects"""
        
        # Convert 'timedelta_series' to type str 
        timedelta_str_series = timedelta_series.astype(str)
        #print(timedelta_str_series)

        # Split 'timedelta_str_series' using the space delimiter and store the results into a list
        timedelta_str_list = [time.split() for time in timedelta_str_series]
        
        # Retrieve the 'time' portion of 'timedelta_str_list' and store into another list  
        time_str_list = [time[2] for time in timedelta_str_list]
    
        # Split 'time_str_list' using '.' delimiter and store results back into 'time_str_list'  
        time_str_list = [time.split('.') for time in time_str_list]

        # Retrieve the '%H:%M:%S' formatted time and store results back into list 
        time_str_list = [time[0] for time in time_str_list]

        # Convert 'time_str_list' into a series and turn each element into a datetime.time() object. 
        # Store in a new list. 
        time_str_series = pd.Series(time_str_list)
        time_obj = [datetime.strptime(time_str, '%H:%M:%S').time() for time_str in time_str_series]
        x_axis = [ datetime.combine(datetime.now(), time) for time in time_obj]
        
        return x_axis

    def make_pdf(self, output_name,  mapping_data_df, create_chart): 
        """Generates a pdf of the processed results 

        Parameters: 
        output_name (str): Name PDF will be saved as 
        mapping_data_df (dataframe): CSV columns to be processed
        create_chart (bool): True if a chart will be generated in the PDF, False if not 
        """
        
        # Get the file path of the wkhtmltopdf executable 
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        df_file = os.getcwd() + '\\' + output_name + '.pdf'
        
        # Replace NaN values with empty strings so the empty data cells do not look like they hold any values in the PDF file 
        mapping_data_df = mapping_data_df.fillna('')
        
        # If the PDF file is to contain a chart, then merge the dataframe and chart PDF into a single PDF. 
        # Otherwise, just save the dataframe PDF as is. 
        if (not create_chart): 
            pdfkit.from_string(mapping_data_df.to_html(), df_file, configuration = config)
    
        else:  
            df_file = os.getcwd() + '\\' + output_name + '_table.pdf'
            pdfkit.from_string(mapping_data_df.to_html(), df_file, configuration = config)
            paths = [os.getcwd() + '\\' + output_name + '_chart.pdf' ,df_file]
            self.merge_pdfs(paths, output_name)

    def merge_pdfs(self,paths, output_name): 
        """Merges two PDFs into a single PDF 
        
        Source: https://realpython.com/pdf-python/

        Parameters: 
        paths (list): File paths of the PDFs to be merged  
        output_name (str): Name PDF will be saved as  
        """
        
        pdf_writer = PdfFileWriter()

        for path in paths: 
            pdf_reader = PdfFileReader(path)
            for page in range(pdf_reader.getNumPages()):
                pdf_writer.addPage(pdf_reader.getPage(page))
        
        with open(output_name + '.pdf', 'wb') as out: 
            pdf_writer.write(out)
        
        # Delete merged files 
        os.remove(paths[0])
        os.remove(paths[1])
    
    def make_txt(self, mapping_df, output_name, format): 
        """Generates a text file of the processed results"""

        mapping_array = mapping_df.to_numpy()
        my_fmt = self.get_format(mapping_df.dtypes, format)

        #TODO: tab delimiter looks weird on txt file
        np.savetxt(output_name + '.txt', mapping_array, fmt = my_fmt, delimiter='\t', header = '\t'.join([str(column) for column in mapping_df.columns]), comments='')


    def get_format(self,dtypes, format): 
        fmt = []
        for i in range(len(dtypes)):
            type = dtypes[i] 
            if (type == np.int64): 
                fmt.append('%d')

            # Parse floats as strings because %f truncates the length of the (very long!) floats 
            else: 
                fmt.append('%s')
        return fmt
    @property
    def get_config_file(self): 
        return self.config_file
    @get_config_file.setter
    def set_config_file(self, config_file): 
        self.config_file = config_file

    @property
    def get_input_csv(self): 
        return self.input_csv
    @get_input_csv.setter
    def set_input_csv(self, input_csv): 
        self.input_csv = input_csv 
    
    #TODO: get_output_name returns an object, not a str
    @property
    def get_output_name(self): 
        return self.output_name 
    @get_output_name.setter
    def set_output_name(self, output_name): 
        self.output_name = output_name

    

